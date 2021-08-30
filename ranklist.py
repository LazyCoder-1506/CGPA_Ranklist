from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ori_wb = load_workbook('original.xlsx')
ori_ws = ori_wb.active

rank_wb = Workbook()

branch = 'AA'

rlist = []

for row in range(1, 1367):
  rollno = ori_ws['A' + str(row)].value
  dep = rollno[2:4]
  student = ori_ws['B' + str(row)].value
  cg = float((ori_ws['C' + str(row)].value).split(' ')[0])
  data = [cg, rollno, student]
  if not dep == branch:
    rlist.sort(reverse=True)
    rank_wb.create_sheet(branch)
    thisdep = rank_wb[branch]
    thisdep.append(['Rank', 'Roll No.', 'Name', 'CGPA'])
    for rank in range(1, len(rlist) + 1):
      data = [rank, rlist[rank-1][1], rlist[rank-1][2], rlist[rank-1][0]]
      thisdep.append(data)

    branch = dep
    rlist = []
  
  record = [cg, rollno, student]
  rlist.append(record)

rank_wb.save('Dep-RankList.xlsx')
print('Done')